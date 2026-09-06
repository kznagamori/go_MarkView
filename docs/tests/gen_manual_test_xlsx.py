#!/usr/bin/env python3
"""docs/specs/41-e2e-manual.md から手動テストの記録用 Excel を生成する。

テストケースの定義は 41 章が正であり（E2E-200）、本スクリプトはそこから
実施記録用のワークブックを組み立てるだけである。ケースを追加・変更する
場合は 41 章を直し、本スクリプトを実行し直す。

使い方:

    python docs/tests/gen_manual_test_xlsx.py --version v1.0.0-rc.1

    -> docs/tests/results/MarkView_手動テスト結果_v1.0.0-rc.1.xlsx

必要なもの: Python 3.10 以上、openpyxl（pip install openpyxl）
Excel のインストールは不要で、Windows / Linux のどちらでも動く。
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.properties import PageSetupProperties
except ImportError:  # pragma: no cover
    sys.exit("openpyxl が見つかりません。`pip install openpyxl` を実行してください。")


# --- 既定のパス（本スクリプトの位置から解決する） ---------------------------

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent.parent
DEFAULT_SRC = REPO_ROOT / "docs" / "specs" / "41-e2e-manual.md"
DEFAULT_OUT_DIR = HERE / "results"


# --- Markdown のパース -------------------------------------------------------

# 節見出し。**節番号を当てにしない**（E2E-200）。グループを増やすと後続の節
# 番号がずれるため、番号で探すと「ケースを 1 つ足しただけで生成が止まる」。
# 番号が付いていてもいなくても、見出しの文言だけを取り出す。
RE_SECTION_HEADING = re.compile(r"^##\s+(?:\d+(?:\.\d+)*\s+)?(.+?)\s*$")
# グループの節は見出しの `Gn` で見分ける（E2E-200）。
RE_GROUP_TITLE = re.compile(r"^G\d+\s+.+$")
# 一覧の節は見出しの文言で探す（E2E-200）。
INDEX_HEADING = "テストケース一覧"
RE_CASE_HEADING = re.compile(r"^###\s+(E2E-\d{3}):\s*(.+?)\s*$")
RE_FIELD = re.compile(r"^-\s+\*\*(環境|優先度|関連要求|概要|前提条件)\*\*:\s*(.+?)\s*$")
RE_BLOCK = re.compile(r"^-\s+\*\*(手順|確認内容)\*\*:\s*$")
# 手順と確認内容は**どちらも番号付きリスト**である（E2E-200）。番号で指摘できる
# ようにするための書式であり、記録用 Excel にも同じ番号で出す。
RE_STEP = re.compile(r"^\s+\d+\.\s+(.+?)\s*$")
# 「テストケース一覧」節の行: | E2E-211 | G1 | 概要 | 高 |
RE_INDEX_ROW = re.compile(r"^\|\s*(E2E-\d{3})\s*\|\s*(G\d+)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*$")

FIELD_ATTR = {
    "環境": "env",
    "優先度": "priority",
    "関連要求": "requirements",
    "概要": "summary",
    "前提条件": "precondition",
}


def section_title(line: str) -> str | None:
    """`## 41.3 G1 起動と終了` から `G1 起動と終了` を取り出す。

    節番号が付いていてもいなくても同じ結果を返す。**番号は当てにしない**
    （E2E-200）。`##` の見出しでなければ None を返す。
    """
    m = RE_SECTION_HEADING.match(line)

    return m.group(1) if m else None


@dataclass
class Case:
    id: str
    title: str
    group: str
    env: str = ""
    priority: str = ""
    requirements: str = ""
    summary: str = ""
    precondition: str = ""
    steps: list[str] = field(default_factory=list)
    expectations: list[str] = field(default_factory=list)

    def missing_fields(self) -> list[str]:
        missing = []
        for label, attr in (
            ("環境", "env"),
            ("優先度", "priority"),
            ("関連要求", "requirements"),
            ("概要", "summary"),
            ("前提条件", "precondition"),
        ):
            if not getattr(self, attr):
                missing.append(label)
        if not self.steps:
            missing.append("手順")
        if not self.expectations:
            missing.append("確認内容")
        return missing


def parse_cases(lines: list[str]) -> list[Case]:
    """グループの節のケース定義を読み取る。

    **グループの見出しは `Gn` で見分ける**（E2E-200）。その下にあるものだけを
    ケースとして扱い、それ以外の節（E2E-200 系の方針や要求一覧）からは
    記録用の行を作らない。**節番号は見ない。**
    """
    cases: list[Case] = []
    group = ""
    current: Case | None = None
    block = ""

    for raw in lines:
        line = raw.rstrip("\n")

        title = section_title(line)
        if title is not None:
            current, block = None, ""
            group = title if RE_GROUP_TITLE.match(title) else ""
            continue

        m = RE_CASE_HEADING.match(line)
        if m:
            block = ""
            if group:
                current = Case(id=m.group(1), title=m.group(2), group=group)
                cases.append(current)
            else:
                current = None
            continue

        if current is None:
            continue

        m = RE_FIELD.match(line)
        if m:
            setattr(current, FIELD_ATTR[m.group(1)], m.group(2))
            block = ""
            continue

        m = RE_BLOCK.match(line)
        if m:
            block = m.group(1)
            continue

        if block == "手順":
            m = RE_STEP.match(line)
            if m:
                current.steps.append(m.group(1))
                continue
        elif block == "確認内容":
            m = RE_STEP.match(line)
            if m:
                current.expectations.append(m.group(1))
                continue

    return cases


def parse_index(lines: list[str]) -> list[tuple[str, str, str, str]]:
    """「テストケース一覧」節のケース一覧を読み取る（突き合わせ用）。

    **節番号ではなく見出しの文言で探す**（E2E-200）。
    """
    rows = []
    in_index = False

    for raw in lines:
        line = raw.rstrip("\n")

        title = section_title(line)
        if title is not None:
            # 一覧の節に入ったあと、次の見出しが来たらそこで終わる。
            if in_index:
                break
            in_index = title == INDEX_HEADING
            continue

        if in_index:
            m = RE_INDEX_ROW.match(line)
            if m:
                rows.append(m.groups())

    return rows


def verify(cases: list[Case], index: list[tuple[str, str, str, str]]) -> list[str]:
    """ケース定義と「テストケース一覧」節を突き合わせ、食い違いを返す。"""
    problems: list[str] = []

    for case in cases:
        missing = case.missing_fields()
        if missing:
            problems.append(f"{case.id}: 項目が欠けている（{', '.join(missing)}）")

    ids = [c.id for c in cases]
    duplicates = sorted({i for i in ids if ids.count(i) > 1})
    if duplicates:
        problems.append(f"ID が重複している: {', '.join(duplicates)}")

    if not index:
        problems.append(f"「{INDEX_HEADING}」節を読み取れなかった")
        return problems

    defined, listed = set(ids), {row[0] for row in index}
    for missing_id in sorted(listed - defined):
        problems.append(f"{missing_id}: 「{INDEX_HEADING}」節にあるが、ケース定義がない")
    for extra_id in sorted(defined - listed):
        problems.append(f"{extra_id}: ケース定義はあるが、「{INDEX_HEADING}」節にない")

    by_id = {c.id: c for c in cases}
    for case_id, group, summary, priority in index:
        case = by_id.get(case_id)
        if case is None:
            continue
        if not case.group.startswith(group + " "):
            problems.append(
                f"{case_id}: グループが一致しない（定義 '{case.group}' / 一覧 '{group}'）"
            )
        if case.priority != priority:
            problems.append(
                f"{case_id}: 優先度が一致しない（定義 '{case.priority}' / 一覧 '{priority}'）"
            )
        if strip_markup(case.title) != strip_markup(summary):
            problems.append(
                f"{case_id}: 概要が一致しない（定義 '{case.title}' / 一覧 '{summary}'）"
            )

    return problems


# --- 書き出し ---------------------------------------------------------------

RE_BOLD = re.compile(r"\*\*(.+?)\*\*")
RE_CODE = re.compile(r"`(.+?)`")


def strip_markup(text: str) -> str:
    """セルに入れるため、強調とコード記法の記号を落とす。"""
    return RE_CODE.sub(r"\1", RE_BOLD.sub(r"\1", text)).strip()


HEADERS = [
    ("ID", 10),
    ("テストグループ", 20),
    ("テスト環境", 12),
    ("優先度", 8),
    ("テスト概要", 40),
    ("前提条件", 32),
    ("テスト手順", 48),
    ("テスト確認内容", 48),
    ("関連要求", 24),
    ("テスト実施日", 14),
    ("ソフトウェアバージョン", 20),
    ("実施環境", 10),
    ("実施者", 10),
    ("テスト確認結果", 14),
    ("実測・備考", 36),
    ("不具合番号", 12),
]
COL_RESULT = 14  # テスト確認結果
RESULTS = ["OK", "NG", "対象外", "未実施"]

HEADER_FILL = PatternFill("solid", fgColor="D9E7F5")
THIN = Side(style="thin", color="9EA7B3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)


def repo_relative(path: Path) -> str:
    """コミットする Excel に絶対パスを残さないよう、リポジトリ相対で表す。"""
    try:
        return path.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return path.name


def build_cover(ws, cases: list[Case], version: str, src: Path) -> None:
    rows = [
        ("MarkView 手動テスト仕様書 兼 結果記録", ""),
        ("", ""),
        ("対象プロダクト", "MarkView"),
        ("リポジトリ", "https://github.com/kznagamori/go_MarkView"),
        ("生成元", f"{repo_relative(src)}（こちらが正）"),
        ("総ケース数", str(len(cases))),
        ("", ""),
        ("■ 実施情報", ""),
        ("ソフトウェアバージョン", version),
        ("実施期間（開始）", ""),
        ("実施期間（終了）", ""),
        ("実施者", ""),
        ("", ""),
        ("■ 結果サマリ（実施時に記入）", ""),
        ("OK", ""),
        ("NG", ""),
        ("対象外", ""),
        ("未実施", ""),
        ("", ""),
        ("■ 注意", ""),
        ("1", "本ファイルは 41 章（E2E テスト仕様: 手動テスト）から生成した記録用ファイルです。"),
        ("2", "ケースの追加・変更は 41 章で行い、本ファイルを生成し直してください。"),
        ("3", "本ファイルを直接編集してケースを変更しないでください。"),
        ("4", "記入するのは J・L〜P 列（実施日・実施環境・実施者・結果・備考・不具合番号）です。"),
        ("5", "実施を終えたら docs/tests/results/ に置いたままコミットしてください。"),
    ]
    for label, value in rows:
        ws.append([label, value])

    ws["A1"].font = Font(size=16, bold=True)
    for row in (8, 14, 20):
        ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=9, column=2).font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72


def build_env(ws) -> None:
    rows = [
        ("コード", "環境", "手動テストでの位置づけ"),
        ("W1", "Windows 11 / amd64", "主環境。全ケースを実施する"),
        ("L1", "Ubuntu 24.04 / amd64", "OS 差の出るケースを実施する"),
        ("WA", "Windows 11 / arm64", "入手できた場合に起動確認を行う"),
        ("LA", "Ubuntu 24.04 / arm64", "入手できた場合に起動確認を行う"),
        ("", "", ""),
        ("優先度", "意味", ""),
        ("高", "壊れていると配布物として成立しない。NG ならリリースしない", ""),
        ("中", "主要機能。NG なら原則リリースしない", ""),
        ("低", "補助的な振る舞い。NG でも影響を判断のうえリリースしてよい", ""),
        ("", "", ""),
        ("結果", "意味", ""),
        ("OK", "確認内容をすべて満たした", ""),
        ("NG", "満たさない項目があった。不具合番号を記入する", ""),
        ("対象外", "環境が用意できない等で実施しなかった。理由を備考に記入する", ""),
        ("未実施", "まだ実施していない", ""),
    ]
    for row in rows:
        ws.append(list(row))
    for row in (1, 7, 12):
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(bold=True)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 40


def build_cases(ws, cases: list[Case], version: str) -> None:
    ws.append([name for name, _ in HEADERS])
    for case in cases:
        ws.append([
            case.id,
            case.group,
            case.env,
            case.priority,
            strip_markup(case.summary),
            strip_markup(case.precondition),
            "\n".join(f"{i}. {strip_markup(s)}" for i, s in enumerate(case.steps, 1)),
            "\n".join(f"{i}. {strip_markup(e)}" for i, e in enumerate(case.expectations, 1)),
            case.requirements,
            "",        # テスト実施日
            version,   # ソフトウェアバージョン
            "",        # 実施環境
            "",        # 実施者
            "未実施",  # テスト確認結果
            "",        # 実測・備考
            "",        # 不具合番号
        ])

    last_row = ws.max_row
    last_col = len(HEADERS)

    for col, (_, width) in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for col in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    for row in range(2, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = TOP_WRAP
            cell.border = BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"

    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(RESULTS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(validation)
    letter = get_column_letter(COL_RESULT)
    validation.add(f"{letter}2:{letter}{last_row}")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"


# --- エントリポイント -------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="41 章から手動テストの記録用 Excel を生成する（E2E-200）。"
    )
    parser.add_argument(
        "--version",
        required=True,
        help="検証する成果物のバージョン（例: v1.0.0-rc.1）。ファイル名と各行に入る",
    )
    parser.add_argument("--src", type=Path, default=DEFAULT_SRC, help="入力の Markdown")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR, help="出力先ディレクトリ")
    parser.add_argument(
        "--force", action="store_true", help="出力先に同名のファイルがあっても上書きする"
    )
    args = parser.parse_args(argv)

    # ファイル名の一部になるため、タグとして妥当な文字だけを受け付ける
    if not re.fullmatch(r"v[0-9A-Za-z.\-+]+", args.version):
        print(
            f"バージョンの形式が不正です: {args.version}"
            "（`v1.0.0` や `v1.0.0-rc.1` の形で指定してください。BR-080）",
            file=sys.stderr,
        )
        return 1

    if not args.src.is_file():
        print(f"入力が見つかりません: {args.src}", file=sys.stderr)
        return 1

    lines = args.src.read_text(encoding="utf-8").splitlines()
    cases = parse_cases(lines)
    if not cases:
        print("ケースを 1 件も抽出できませんでした。41 章の書式を確認してください。", file=sys.stderr)
        return 1

    problems = verify(cases, parse_index(lines))
    if problems:
        print("41 章の内容に食い違いがあります。ファイルは作成しません。", file=sys.stderr)
        for problem in problems:
            print(f"  - {problem}", file=sys.stderr)
        return 1

    out_path = args.out_dir / f"MarkView_手動テスト結果_{args.version}.xlsx"
    if out_path.exists() and not args.force:
        print(f"すでに存在します: {out_path}", file=sys.stderr)
        print("実施済みの記録を消さないよう、上書きするなら --force を付けてください。", file=sys.stderr)
        return 1

    book = Workbook()
    cover = book.active
    cover.title = "表紙"
    build_cover(cover, cases, args.version, args.src)
    build_cases(book.create_sheet("テスト仕様・結果"), cases, args.version)
    build_env(book.create_sheet("環境定義"))

    args.out_dir.mkdir(parents=True, exist_ok=True)
    book.save(out_path)

    print(f"生成しました: {out_path}")
    print(f"  ケース数: {len(cases)}  バージョン: {args.version}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
